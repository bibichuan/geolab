#!/usr/bin/python
# -*- coding: UTF-8 -*-
import openpyxl as excel

## 读取excel,获取站点关联线段
excelFile=r"D:\zlc\test.xlsx"
wb=excel.load_workbook(excelFile)
## 读取第一行，获取列头
sheet=wb["Sheet1"]
rows=sheet.max_row
columns=sheet.max_column
array=[]

test="kkj就立即立即"
for kl in test.split("立"):
    print kl

for i in range(1,rows-1):
    temp=sheet.cell(row=i,column=1).value
    ## 分割
    # temp=temp.encode("utf-8")
    # print temp
    # sp="″".encode("utf-8")
    coordinates=temp.encode('utf-8').split("″")
    # for c in coordinates:
    #     print c
    print coordinates
    print len(coordinates)
    # for j in range(len(coordinates)-2):
    #     print j
    #     te=coordinates[j]
    #     print coordinates.get
        # if temp not None:
        #     print temp   
    # print coordinates
    # array.append(temp)