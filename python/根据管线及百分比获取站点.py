#!/usr/bin/python
# -*- coding: UTF-8 -*-

import os
import arcpy
import openpyxl as excel

## 读取excel,获取站点关联线段
excelFile=r"D:\zlc\hdxs\filedata\station_20190826_1701.xlsx"
wb=excel.load_workbook(excelFile)
## 读取第一行，获取列头
sheet=wb["Sheet1"]
rows=sheet.max_row
columns=sheet.max_column
dic={}
for i in range(1,rows-1):
    dic[sheet.cell(row=1,column=i).value]=i

shp = r'D:\zlc\hdxs\geodata\pipeline_20190819_1218.shp'
for i in range(2,rows+1):
    id=sheet.cell(i,dic['id']).value
    if id is None:
        continue
    lineid=sheet.cell(i,dic['lineid']).value
    percentage=sheet.cell(i,dic['percentage']).value
    ## 查询图层
    row=None
    with arcpy.da.SearchCursor(shp,["Number","StartLon","StartLat","EndLon","EndLat"],"Number = "+str(lineid)) as cursor:
        for r in cursor:
            row=r
    if row is not None:
        startLon=row[1]
        startLat=row[2]
        endLon=row[3]
        endLat=row[4]
        tempLon=round(float(startLon)+(float(endLon)-float(startLon))*float(percentage),6)
        tempLat=round(float(startLat)+(float(endLat)-float(startLat))*float(percentage),6)
        print tempLon
        print tempLat

    ## 修改经纬度坐标值
    sheet.cell(i,dic['x']).value=tempLon
    sheet.cell(i,dic['y']).value=tempLat

    print str(tempLat)+"_"+str(tempLon)

wb.save(filename=excelFile)
print "done"