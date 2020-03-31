#!/usr/bin/python
# -*- coding: UTF-8 -*-

import math
import openpyxl as excel
import arcpy
import os

## 大地坐标系资料WGS-84 长半径a=6378137 短半径b=6356752.3142 扁率f=1/298.2572236

## 长半径a=6378137 
a = 6378137
## 短半径b=6356752.3142
b = 6356752.3142
## 扁率f=1/298.2572236
f = 1 / 298.2572236

## lon 经度，lat 纬度，brng 方位角，dist 距离(m)
def computerThatLonLat(lon,lat,brng,dist):
    alpha1 = rad(brng)
    sinAlpha1 = math.sin(alpha1)
    cosAlpha1 = math.cos(alpha1)

    tanU1 = (1 - f) * math.tan(rad(lat))
    cosU1 = 1 / math.sqrt((1 + tanU1 * tanU1))
    sinU1 = tanU1 * cosU1
    sigma1 = math.atan2(tanU1, cosAlpha1)
    sinAlpha = cosU1 * sinAlpha1
    cosSqAlpha = 1 - sinAlpha * sinAlpha
    uSq = cosSqAlpha * (a * a - b * b) / (b * b)
    A = 1 + uSq / 16384 * (4096 + uSq * (-768 + uSq * (320 - 175 * uSq)))
    B = uSq / 1024 * (256 + uSq * (-128 + uSq * (74 - 47 * uSq)))

    cos2SigmaM=0
    sinSigma=0
    cosSigma=0
    sigma = dist / (b * A)
    sigmaP = 2 * math.pi
    while math.fabs(sigma - sigmaP) > 1e-12:
        cos2SigmaM = math.cos(2 * sigma1 + sigma)
        sinSigma = math.sin(sigma)
        cosSigma = math.cos(sigma)
        deltaSigma = B * sinSigma * (cos2SigmaM + B / 4 * (cosSigma * (-1 + 2 * cos2SigmaM * cos2SigmaM)
                - B / 6 * cos2SigmaM * (-3 + 4 * sinSigma * sinSigma) * (-3 + 4 * cos2SigmaM * cos2SigmaM)))
        sigmaP = sigma
        sigma = dist / (b * A) + deltaSigma
    
    tmp = sinU1 * sinSigma - cosU1 * cosSigma * cosAlpha1
    lat2 = math.atan2(sinU1 * cosSigma + cosU1 * sinSigma * cosAlpha1, (1 - f) * math.sqrt(sinAlpha * sinAlpha + tmp * tmp))
    lamb = math.atan2(sinSigma * sinAlpha1, cosU1 * cosSigma - sinU1 * sinSigma * cosAlpha1)
    C = f / 16 * cosSqAlpha * (4 + f * (4 - 3 * cosSqAlpha))
    L = lamb - (1 - C) * f * sinAlpha * (sigma + C * sinSigma * (cos2SigmaM + C * cosSigma * (-1 + 2 * cos2SigmaM * cos2SigmaM)))

    lon=lon+deg(L)
    lat=deg(lat2)

    return [lon,lat]

## 度换成弧度,d 弧度
def rad(d):
    return d * math.pi / 180.0
## 弧度换成度,x 弧度
def deg(x):
    return x * 180 / math.pi

## 读取excel，根据excel表格生成对应的坐标数据
def handleExecl(path):
    ## 读取excel,获取站点关联线段
    excelFile=path
    wb=excel.load_workbook(excelFile)
    ## 读取第一行，获取列头
    sheet=wb["Sheet1"]
    rows=sheet.max_row
    columns=sheet.max_column
    dic={}
    ## 存储列名与列号的对应
    for i in range(1,columns+1):
        dic[sheet.cell(row=1,column=i).value]=i

    ## 遍历剩余的行，根据对应关系求坐标

    pointlist=[]
    for i in range(2,rows+1):
        start_n=str(sheet.cell(i,dic['start_n']).value)

        if start_n is None:
            continue

        end_n=str(sheet.cell(i,dic['end_n']).value)
        angle=sheet.cell(i,dic['angle']).value
        dis=sheet.cell(i,dic['dis']).value
        start_lon=sheet.cell(i,dic['start_lon']).value
        start_lat=sheet.cell(i,dic['start_lat']).value
        end_lon=sheet.cell(i,dic['end_lon']).value
        end_lat=sheet.cell(i,dic['end_lat']).value

        temp={'start_n':start_n,'end_n':end_n,'angle':angle,'dis':dis,'start_lon':start_lon,'start_lat':start_lat,'end_lon':end_lon,'end_lat':end_lat}
    
        ## 计算经纬度
        ## 如果开始经纬度存在，则直接计算终点经纬度，否则，先根据其前一个点找个经纬度，然后计算终点
        if (not start_lat) and (not start_lon):
            for item in pointlist:
                if str(item['start_n']) == start_n:
                    start_lat=item['start_lat']
                    start_lon=item['start_lon']
                    break
                if item['end_n'] == start_n:
                    start_lat=item['end_lat']
                    start_lon=item['end_lon']
                    break
        temp['start_lat']=start_lat
        temp['start_lon']=start_lon
        sheet.cell(i,dic['start_lat']).value=start_lat
        sheet.cell(i,dic['start_lon']).value=start_lon
        if start_lat and start_lon:
            result=computerThatLonLat(start_lon,start_lat,angle,dis)

            temp['end_lon']=result[0]
            temp['end_lat']=result[1]

            ## 保存到数据表中
            sheet.cell(i,dic['end_lon']).value=result[0]
            sheet.cell(i,dic['end_lat']).value=result[1]

        pointlist.append(temp)
            
        # temp={'start_n':start_n,'end_n':end_n,'angle':angle,'dis':dis,'start_lon':start_lon,'start_lat':start_lat,'end_lon':end_lon,'end_lat':end_lat}
        # pointDic[start_n]=temp
        # ## 修改经纬度坐标值
        # sheet.cell(i,dic['x']).value=tempLon
        # sheet.cell(i,dic['y']).value=tempLat

        # print str(tempLat)+"_"+str(tempLon)
    wb.save(filename=excelFile)
    print "done"

## 读取excel生成shp文件
def excelToShp(path):

    currentPath=os.getcwd()
    arcpy.env.workspace =currentPath
    if not os.path.isdir(currentPath + '/shp'):
        os.mkdir(currentPath + '/shp')
    outName =os.path.join(os.getcwd(), 'shp', 'polyline.shp') 
    polylineGeometryList = [] # a list to hold the PointGeometry objects
    point = arcpy.Point() #create an empty Point object
    spRef = arcpy.SpatialReference(4326)


    ## 读取excel,获取站点关联线段
    excelFile=path
    wb=excel.load_workbook(excelFile)
    ## 读取第一行，获取列头
    sheet=wb["Sheet1"]
    rows=sheet.max_row
    columns=sheet.max_column
    dic={}
    ## 存储列名与列号的对应
    for i in range(1,columns+1):
        dic[sheet.cell(row=1,column=i).value]=i

    ## 遍历剩余的行，根据对应关系求坐标
    polylinGeometryList=[]
    for i in range(2,rows+1):
        start_n=str(sheet.cell(i,dic['start_n']).value)

        if start_n is None:
            continue
        
        ## 开始坐标和结束坐标
        start_lon=sheet.cell(i,dic['start_lon']).value
        start_lat=sheet.cell(i,dic['start_lat']).value
        end_lon=sheet.cell(i,dic['end_lon']).value
        end_lat=sheet.cell(i,dic['end_lat']).value
        ## 构建线段
        startPoint=arcpy.Point(float(start_lon), float(start_lat))
        endPoint=arcpy.Point(float(end_lon), float(end_lat))
        array=arcpy.Array()
        array.append(startPoint)
        array.append(endPoint)
        polylinGeometry = arcpy.Polyline(array,spRef)
        polylinGeometryList.append(polylinGeometry)
    
    ## 保存内容
    arcpy.CopyFeatures_management(polylinGeometryList, outName)

    print 'done'



if __name__=='__main__':

    # lon=120.353842
    # lat=30.338461
    # brng=90
    # dist=1
    # computerThatLonLat(lon, lat, brng, dist)

    excelPath=r"D:\zlc\hdxs\filedata\creat_test.xlsx"
    ## handleExecl(excelPath)
    excelToShp(excelPath)
    